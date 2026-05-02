"""XLSX report generator — openpyxl.

Layout:
    Sheet 1 — Summary  (one row per analysed combination)
    Sheet 2…N — one sheet per combination with the re-gridded ref/test
                arrays + diff column so APTRANSCO reviewers can plot
                directly in Excel.
    Sheet _metadata — full session metadata.
"""
from __future__ import annotations

import io
from typing import Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from sfra_full import __version__
from sfra_full.db import (
    AnalysisResult,
    Combination,
    OverhaulCycle,
    TestSession,
    Trace,
    Transformer,
)
from sfra_full.db.array_helpers import bytes_to_array


_VERDICT_FILL = {
    "NORMAL": "10b981",
    "APPEARS_NORMAL": "10b981",
    "MINOR_DEVIATION": "f59e0b",
    "SUSPECT": "f59e0b",
    "SIGNIFICANT_DEVIATION": "f97316",
    "SEVERE_DEVIATION": "f43f5e",
    "INDETERMINATE": "94a3b8",
}


def build_session_xlsx(
    *,
    transformer: Transformer,
    cycle: OverhaulCycle,
    session: TestSession,
    analyses: list[AnalysisResult],
    combinations: list[Combination],
    traces_by_id: dict[str, Trace],
    expected_total: int,
) -> bytes:
    """Render the session XLSX and return its bytes."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    summary_ws.append([
        "Combination", "Mode", "Severity", "CC_LOW", "CC_MID_L", "CC_MID",
        "CC_HIGH", "RL_LOW", "RL_MID_L", "RL_MID", "RL_HIGH",
        "n_matched", "n_lost", "n_new", "Auto-remark",
    ])
    for col in summary_ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill("solid", fgColor="1E3A8A")
        col.alignment = Alignment(horizontal="center")

    code_to_combo = {c.id: c for c in combinations}
    for ar in sorted(analyses, key=lambda a: (a.combination_id or 0)):
        combo = code_to_combo.get(ar.combination_id) if ar.combination_id else None
        ind = ar.indicators_json or {}
        per_band = {
            (b.get("band_code") or ""): b for b in (ind.get("per_band") or [])
        }
        row = [
            combo.code if combo else "—",
            ar.mode.value,
            ar.severity.value,
            _safe(per_band.get("LOW", {}).get("cc")),
            _safe(per_band.get("MID_L", {}).get("cc")),
            _safe(per_band.get("MID", {}).get("cc")),
            _safe(per_band.get("HIGH", {}).get("cc")),
            _safe(per_band.get("LOW", {}).get("rl_factor")),
            _safe(per_band.get("MID_L", {}).get("rl_factor")),
            _safe(per_band.get("MID", {}).get("rl_factor")),
            _safe(per_band.get("HIGH", {}).get("rl_factor")),
            ind.get("n_matched"),
            ind.get("n_lost"),
            ind.get("n_new"),
            (ar.auto_remarks or "")[:200],
        ]
        summary_ws.append(row)
        new_row = summary_ws.max_row
        fill = _VERDICT_FILL.get(ar.severity.value)
        if fill:
            summary_ws.cell(row=new_row, column=3).fill = PatternFill(
                "solid", fgColor=fill
            )
            summary_ws.cell(row=new_row, column=3).font = Font(
                bold=True, color="FFFFFF"
            )

    # ---------- Per-combination sheets ----------
    for ar in analyses:
        combo = code_to_combo.get(ar.combination_id) if ar.combination_id else None
        sheet_name = (combo.code if combo else f"trace-{ar.tested_trace_id[:6]}")[:31]
        ws = wb.create_sheet(sheet_name)
        tested = traces_by_id.get(ar.tested_trace_id)
        ref = traces_by_id.get(ar.reference_trace_id) if ar.reference_trace_id else None
        if tested is None:
            ws.append(["No tested trace data."])
            continue

        f_test = bytes_to_array(tested.frequency_hz)
        m_test = bytes_to_array(tested.magnitude_db)
        p_test = bytes_to_array(tested.phase_deg)
        f_ref = bytes_to_array(ref.frequency_hz) if ref else None
        m_ref = bytes_to_array(ref.magnitude_db) if ref else None

        if ref and f_ref is not None and m_ref is not None and f_test is not None and m_test is not None:
            # Interpolate the reference onto the tested grid for diff column.
            order = np.argsort(f_ref)
            m_ref_on_test = np.interp(f_test, f_ref[order], m_ref[order])
            ws.append(["frequency_hz", "ref_mag_db", "test_mag_db", "diff_db", "test_phase_deg"])
            for fi, mi_ref, mi_test, ph in zip(
                f_test, m_ref_on_test, m_test, (p_test if p_test is not None else [None] * len(f_test)),
                strict=False,
            ):
                ws.append([float(fi), float(mi_ref), float(mi_test), float(mi_test - mi_ref),
                           float(ph) if ph is not None else None])
        else:
            ws.append(["frequency_hz", "test_mag_db", "test_phase_deg"])
            for fi, mi, ph in zip(f_test, m_test, (p_test if p_test is not None else [None] * len(f_test)), strict=False):
                ws.append([float(fi), float(mi), float(ph) if ph is not None else None])

        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill("solid", fgColor="1E3A8A")

    # ---------- Metadata sheet ----------
    meta = wb.create_sheet("_metadata")
    meta.append(["Field", "Value"])
    meta.append(["Engine version", __version__])
    meta.append(["Transformer serial", transformer.serial_no])
    meta.append(["Type", transformer.transformer_type.value])
    meta.append(["Vector group", transformer.vector_group or ""])
    meta.append(["MVA", transformer.nameplate_mva or ""])
    meta.append(["HV kV", transformer.hv_kv or ""])
    meta.append(["LV kV", transformer.lv_kv or ""])
    meta.append(["Substation", transformer.substation or ""])
    meta.append(["Cycle no", cycle.cycle_no])
    meta.append(["Cycle start", str(cycle.cycle_start_date)])
    meta.append(["Cycle end", str(cycle.cycle_end_date) if cycle.cycle_end_date else ""])
    meta.append(["Session type", session.session_type.value])
    meta.append(["Session date", str(session.session_date)])
    meta.append(["Tested by", session.tested_by or ""])
    meta.append(["Instrument", session.instrument_make_model or ""])
    meta.append(["Ambient °C", session.ambient_temp_c or ""])
    meta.append(["Oil °C", session.oil_temp_c or ""])
    meta.append(["Analyses present", len(analyses)])
    meta.append(["Catalogue total", expected_total])
    meta.append(["Is partial", len(analyses) < expected_total])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


__all__ = ["build_session_xlsx"]
